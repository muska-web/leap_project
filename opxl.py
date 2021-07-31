import openpyxl as op
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import glob as gb
import sqlite3 as sql3
from datetime import datetime as dt

#db connection
dbName = 'Data'
conn = sql3.connect(dbName + '.sqlite')
cur = conn.cursor()

columnsToBeExtractedFromDataSheet = ['Date', 'FacilityType', 'BedSize', 'Region', 'Manufacturer', 'Ticker', 'Group', 'Therapy', 'Anatomy','SubAnatomy', 'ProductCategory', 'Quantity', 'AvgPrice', 'TotalSpend']

curDate = dt.today()
outputFileName = "model_" + curDate.strftime("%Y%m%d")
print(outputFileName)
outputWorkbookPath = "C:\\Users\\janam\\Desktop\\ETL Datasheet\\"+outputFileName+".xlsx"
outputWorkbook = Workbook()
outputWorkbookWorksheet = outputWorkbook.active
outputWorkbookWorksheet['A1'] = 'Date'
outputWorkbookWorksheet['B1'] = 'Ticker'
outputWorkbookWorksheet['C1'] = 'Type'
outputWorkbookWorksheet['D1'] = 'Quarter'
outputWorkbookWorksheet['E1'] = 'Year'
outputWorkbookWorksheet['F1'] = 'Estimated Total Sold'
outputWorkbookWorksheet['G1'] = 'Estimated Sold Maximum'
outputWorkbookWorksheet['H1'] = 'Estimated Sold Minimum'
outputWorkbookWorksheet['I1'] = 'Forecast w/o SA'
outputWorkbookWorksheet['J1'] = 'Forecase w/o Max'
outputWorkbookWorksheet['K1'] = 'Forecast w/o Min'

data = []
path = os.getcwd()
files = os.listdir(path)

files = [file for file in files if ".xlsx" in file]

#checking if files already exist
#need to change directory paths
curDate = dt.today()
fileName = "model_" + curDate.strftime("%Y%m%d") + ".xlsx"
created = os.stat(fileName).st_ctime
now = dt.now()
dateTimeFile = dt.fromtimestamp(created)

dateFile = dateTimeFile.strftime("%m/%d/%Y")

dateNow = now.strftime("%m/%d/%Y")

if(dateNow == dateFile):
    os.remove(fileName)
    print("fileRemoved")

rowCounterEmpirical = 2
rowCounterRegression = 2

for file in files:
    ticker = file.split(' ')[0]
    workbook = op.load_workbook(r'C:\\Users\\janam\\Desktop\\ETL Datasheet\\'+file, read_only=True, data_only=True)
    allWorksheetsInTheFile = workbook.sheetnames
    print("Loading: " + file)
    empericalModelSheets = [sheets for sheets in allWorksheetsInTheFile if "Empirical Model" in sheets]
    data = [sheets for sheets in allWorksheetsInTheFile if sheets == "Data"]
    regressionModelSheets = [sheets for sheets in allWorksheetsInTheFile if "Regression Model" in sheets]

    #sheets with emperical model data
    for sheet in empericalModelSheets:
        workableSheet = workbook[sheet]
        print(sheet)
        for row in workableSheet['D1':'D' + str(workableSheet.max_row)]:
            for cellValue in row:
                tempStr = str(cellValue.value)
                if "Estimated total sold" in tempStr and tempStr[-3] == "Q":
                    estimatedTotalSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row)]
                    estimatedMaxSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row + 1)]
                    estimatedMinSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row + 2)]

                    print(estimatedTotalSold.value)
                    print(estimatedMaxSold.value)
                    print(estimatedMinSold.value)

                    outputWorkbookWorksheet["F"+str(rowCounterEmpirical)] = estimatedTotalSold.value
                    print("F" + str(rowCounterEmpirical))
                    outputWorkbookWorksheet["G"+str(rowCounterEmpirical)] = estimatedMaxSold.value
                    print("G" + str(rowCounterEmpirical))
                    outputWorkbookWorksheet["H"+str(rowCounterEmpirical)] = estimatedMinSold.value
                    print("H" + str(rowCounterEmpirical))
                    if sheet[-5:] != "Model" :
                        sheetNameTmp = sheet.split('-')
                        print(sheetNameTmp[1].strip())
                        outputWorkbookWorksheet["C"+str(rowCounterEmpirical)] = sheetNameTmp[1].strip()
                    else:
                        outputWorkbookWorksheet["C"+str(rowCounterEmpirical)] = "Null"
                    rowCounterEmpirical = rowCounterEmpirical + 1
    
    #sheets with regression model data
    for sheet in regressionModelSheets:
        workableSheet = workbook[sheet]
        print(sheet)
        for row in workableSheet["C1":"R" + str(workableSheet.max_row)]:
            for colObj in row:
                tmpValue = str(colObj.value)
                if tmpValue.strip() == "Max":
                    quarter = workableSheet["D" + str(colObj.row - 1)].value
                    yeartmp = workableSheet["C" + str(colObj.row -1)].value
                    year = "20" + yeartmp[-2:]
                    forecastSA = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row - 1)].value
                    forecastMin = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row + 1)].value
                    forecastMax = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row)].value
                    outputWorkbookWorksheet["A"+str(rowCounterRegression)] = curDate.strftime("%Y-%m-%d")
                    print("A" + str(rowCounterRegression))
                    outputWorkbookWorksheet["D"+str(rowCounterRegression)] = quarter
                    print("D" + str(rowCounterRegression))
                    outputWorkbookWorksheet["B"+str(rowCounterRegression)] = ticker
                    print("B" + str(rowCounterRegression))
                    outputWorkbookWorksheet["E"+str(rowCounterRegression)] = year
                    print("E" + str(rowCounterRegression))
                    outputWorkbookWorksheet["I"+str(rowCounterRegression)] = forecastSA
                    print("I" + str(rowCounterRegression))
                    outputWorkbookWorksheet["J"+str(rowCounterRegression)] = forecastMax
                    print("J" + str(rowCounterRegression))
                    outputWorkbookWorksheet["K"+str(rowCounterRegression)] = forecastMin
                    print("K" + str(rowCounterRegression))
                    if sheet[-5:] != "Model" :
                        sheetNameTmp = sheet.split('-')
                        print(sheetNameTmp[1].strip())
                        outputWorkbookWorksheet["C"+str(rowCounterRegression)] = sheetNameTmp[1].strip()
                    else:
                        outputWorkbookWorksheet["C"+str(rowCounterRegression)] = "Null"
                    
                    rowCounterRegression = rowCounterRegression + 1

outputWorkbook.save(outputWorkbookPath)

#database storage part
